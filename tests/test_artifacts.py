"""
Tests for ArtifactsService — BQ-VZ-ARTIFACTS Phase 1

Covers: create, list, get, download, delete, star, cleanup,
        quotas, rate limits, path traversal prevention, content validation.
"""

import json
import os
import time
from datetime import datetime, timezone, timedelta
from types import SimpleNamespace

import pytest

# Set test env before importing app modules
os.environ.setdefault("VECTORAIZ_AUTH_ENABLED", "false")
os.environ.setdefault("VECTORAIZ_DEBUG", "true")
os.environ.setdefault("ENVIRONMENT", "development")


@pytest.fixture
def artifacts_dir(tmp_path):
    """Create a temp artifacts directory."""
    d = tmp_path / "artifacts"
    d.mkdir()
    return d


@pytest.fixture
def svc(artifacts_dir):
    """Create an ArtifactsService with a temp data directory."""
    from app.services.artifacts_service import ArtifactsService

    service = ArtifactsService()
    service._artifacts_dir = artifacts_dir
    service._rate_limits.clear()
    return service


class TestCreateArtifact:
    def test_create_basic_txt(self, svc):
        a = svc.create_artifact(
            filename="test.txt",
            content="Hello, world!",
            format="txt",
            description="Test artifact",
            dataset_refs=["ds1"],
            user_id="local",
        )
        assert a.id
        assert a.filename == "test.txt"
        assert a.format == "txt"
        assert a.size_bytes == len("Hello, world!".encode("utf-8"))
        assert a.user_id == "local"
        assert a.starred is False
        assert a.expired is False
        assert a.description == "Test artifact"
        assert a.dataset_refs == ["ds1"]
        assert a.source == "allai-copilot"

    def test_create_csv(self, svc):
        content = "name,age\nAlice,30\nBob,25"
        a = svc.create_artifact(
            filename="data.csv", content=content, format="csv",
            description="CSV test", dataset_refs=[], user_id="local",
        )
        assert a.format == "csv"
        assert a.size_bytes > 0

    def test_create_json(self, svc):
        content = '{"key": "value"}'
        a = svc.create_artifact(
            filename="data.json", content=content, format="json",
            description="JSON test", dataset_refs=[], user_id="local",
        )
        assert a.format == "json"

    def test_create_md(self, svc):
        content = "# Title\n\nSome markdown content."
        a = svc.create_artifact(
            filename="report.md", content=content, format="md",
            description="MD test", dataset_refs=[], user_id="local",
        )
        assert a.format == "md"

    def test_create_html_strips_scripts(self, svc):
        content = '<html><body><script>alert("xss")</script><p>Hello</p></body></html>'
        a = svc.create_artifact(
            filename="page.html", content=content, format="html",
            description="HTML test", dataset_refs=[], user_id="local",
        )
        # Read content back
        content_path = svc._get_artifact_dir(a.id) / "content.html"
        saved = content_path.read_text(encoding="utf-8")
        assert "<script" not in saved
        assert "<p>Hello</p>" in saved

    def test_content_file_uses_format_ext_not_filename(self, svc):
        """Content file is content.{ext} from format enum, not user's filename."""
        a = svc.create_artifact(
            filename="my-report.txt", content="test", format="txt",
            description="test", dataset_refs=[], user_id="local",
        )
        artifact_dir = svc._get_artifact_dir(a.id)
        assert (artifact_dir / "content.txt").exists()
        assert not (artifact_dir / "my-report.txt").exists()

    def test_create_with_no_user_id_raises(self, svc):
        with pytest.raises(ValueError, match="user_id is required"):
            svc.create_artifact(
                filename="test.txt", content="hello", format="txt",
                description="test", dataset_refs=[], user_id="",
            )

    def test_create_atomic_writes(self, svc):
        """Artifact dir should exist after create, no temp dirs left."""
        a = svc.create_artifact(
            filename="test.txt", content="content", format="txt",
            description="test", dataset_refs=[], user_id="local",
        )
        artifact_dir = svc._get_artifact_dir(a.id)
        assert artifact_dir.exists()
        assert (artifact_dir / "metadata.json").exists()
        assert (artifact_dir / "content.txt").exists()
        # No temp dirs should remain
        temp_dirs = [d for d in svc._artifacts_dir.iterdir() if d.name.startswith(".tmp_")]
        assert len(temp_dirs) == 0

    def test_content_hash_is_sha256(self, svc):
        import hashlib
        content = "hash test content"
        a = svc.create_artifact(
            filename="hash.txt", content=content, format="txt",
            description="test", dataset_refs=[], user_id="local",
        )
        expected = hashlib.sha256(content.encode("utf-8")).hexdigest()
        assert a.content_hash == expected


class TestValidation:
    def test_empty_filename_rejected(self, svc):
        with pytest.raises(ValueError, match="empty"):
            svc.create_artifact(
                filename="", content="hello", format="txt",
                description="test", dataset_refs=[], user_id="local",
            )

    def test_path_traversal_sanitized(self, svc):
        """Path separators are stripped by os.path.basename — filename becomes 'passwd'."""
        a = svc.create_artifact(
            filename="../../etc/passwd", content="hello", format="txt",
            description="test", dataset_refs=[], user_id="local",
        )
        # basename strips path components — stored as 'passwd'
        assert a.filename == "passwd"
        # The actual file is content.txt, not the filename
        assert (svc._get_artifact_dir(a.id) / "content.txt").exists()

    def test_dotdot_in_filename_rejected(self, svc):
        with pytest.raises(ValueError):
            svc.create_artifact(
                filename="foo..bar.txt", content="hello", format="txt",
                description="test", dataset_refs=[], user_id="local",
            )

    def test_special_chars_rejected(self, svc):
        with pytest.raises(ValueError, match="invalid characters"):
            svc.create_artifact(
                filename="file name.txt", content="hello", format="txt",
                description="test", dataset_refs=[], user_id="local",
            )

    def test_nul_bytes_rejected(self, svc):
        with pytest.raises(ValueError, match="NUL"):
            svc.create_artifact(
                filename="test.txt", content="hello\x00world", format="txt",
                description="test", dataset_refs=[], user_id="local",
            )

    def test_invalid_json_rejected(self, svc):
        with pytest.raises(ValueError, match="Invalid JSON"):
            svc.create_artifact(
                filename="data.json", content="{bad json", format="json",
                description="test", dataset_refs=[], user_id="local",
            )

    def test_invalid_format_rejected(self, svc):
        with pytest.raises(ValueError):
            svc.create_artifact(
                filename="test.exe", content="hello", format="exe",
                description="test", dataset_refs=[], user_id="local",
            )

    def test_hidden_file_rejected(self, svc):
        with pytest.raises(ValueError, match="cannot start with"):
            svc.create_artifact(
                filename=".hidden", content="hello", format="txt",
                description="test", dataset_refs=[], user_id="local",
            )


class TestListGetDownloadDelete:
    def test_list_returns_user_artifacts(self, svc):
        svc.create_artifact("a.txt", "content a", "txt", "a", [], "user1")
        svc.create_artifact("b.txt", "content b", "txt", "b", [], "user2")
        svc.create_artifact("c.txt", "content c", "txt", "c", [], "user1")

        result = svc.list_artifacts("user1")
        assert len(result) == 2
        filenames = {a.filename for a in result}
        assert filenames == {"a.txt", "c.txt"}

    def test_list_sorted_by_created_desc(self, svc):
        svc.create_artifact("first.txt", "1", "txt", "first", [], "local")
        time.sleep(0.01)
        svc.create_artifact("second.txt", "2", "txt", "second", [], "local")

        result = svc.list_artifacts("local")
        assert result[0].filename == "second.txt"
        assert result[1].filename == "first.txt"

    def test_list_pagination(self, svc):
        for i in range(5):
            svc.create_artifact(f"file{i}.txt", f"content{i}", "txt", f"d{i}", [], "local")

        result = svc.list_artifacts("local", offset=2, limit=2)
        assert len(result) == 2

    def test_get_artifact(self, svc):
        a = svc.create_artifact("test.txt", "hello", "txt", "desc", [], "local")
        fetched = svc.get_artifact(a.id, "local")
        assert fetched.id == a.id
        assert fetched.filename == "test.txt"

    def test_get_artifact_wrong_user(self, svc):
        a = svc.create_artifact("test.txt", "hello", "txt", "desc", [], "user1")
        with pytest.raises(FileNotFoundError):
            svc.get_artifact(a.id, "user2")

    def test_download_artifact(self, svc):
        a = svc.create_artifact("test.txt", "hello world", "txt", "desc", [], "local")
        path, filename, mime = svc.download_artifact(a.id, "local")
        assert path.exists()
        assert filename == "test.txt"
        assert mime == "text/plain"
        assert path.read_text(encoding="utf-8") == "hello world"

    def test_delete_artifact(self, svc):
        a = svc.create_artifact("test.txt", "hello", "txt", "desc", [], "local")
        assert svc.delete_artifact(a.id, "local") is True
        with pytest.raises(FileNotFoundError):
            svc.get_artifact(a.id, "local")

    def test_delete_wrong_user(self, svc):
        a = svc.create_artifact("test.txt", "hello", "txt", "desc", [], "user1")
        with pytest.raises(FileNotFoundError):
            svc.delete_artifact(a.id, "user2")


class TestStar:
    def test_star_and_unstar(self, svc):
        a = svc.create_artifact("test.txt", "hello", "txt", "desc", [], "local")
        assert a.starred is False

        updated = svc.star_artifact(a.id, "local", True)
        assert updated.starred is True

        updated2 = svc.star_artifact(a.id, "local", False)
        assert updated2.starred is False

    def test_star_persists(self, svc):
        a = svc.create_artifact("test.txt", "hello", "txt", "desc", [], "local")
        svc.star_artifact(a.id, "local", True)
        fetched = svc.get_artifact(a.id, "local")
        assert fetched.starred is True


class TestCleanup:
    def test_cleanup_removes_old_unstarred(self, svc):
        a = svc.create_artifact("old.txt", "hello", "txt", "desc", [], "local")
        # Backdate the artifact
        meta_path = svc._get_artifact_dir(a.id) / "metadata.json"
        with open(meta_path, "r") as f:
            data = json.load(f)
        old_time = (datetime.now(timezone.utc) - timedelta(days=8)).isoformat()
        data["created_at"] = old_time
        with open(meta_path, "w") as f:
            json.dump(data, f)

        removed = svc.cleanup_expired()
        assert removed == 1
        assert not svc._get_artifact_dir(a.id).exists()

    def test_cleanup_keeps_starred(self, svc):
        a = svc.create_artifact("starred.txt", "hello", "txt", "desc", [], "local")
        svc.star_artifact(a.id, "local", True)
        # Backdate
        meta_path = svc._get_artifact_dir(a.id) / "metadata.json"
        with open(meta_path, "r") as f:
            data = json.load(f)
        old_time = (datetime.now(timezone.utc) - timedelta(days=30)).isoformat()
        data["created_at"] = old_time
        with open(meta_path, "w") as f:
            json.dump(data, f)

        removed = svc.cleanup_expired()
        assert removed == 0
        assert svc._get_artifact_dir(a.id).exists()

    def test_cleanup_keeps_recent(self, svc):
        svc.create_artifact("recent.txt", "hello", "txt", "desc", [], "local")
        removed = svc.cleanup_expired()
        assert removed == 0


class TestQuotas:
    def test_rate_limit_enforced(self, svc):
        for i in range(5):
            svc.create_artifact(f"file{i}.txt", f"content{i}", "txt", "d", [], "local")
        with pytest.raises(PermissionError, match="Rate limit"):
            svc.create_artifact("file5.txt", "content5", "txt", "d", [], "local")

    def test_count_quota_enforced(self, svc):
        # Create 100 artifacts (need to bypass rate limit)
        for i in range(100):
            svc._rate_limits.clear()  # bypass rate limit for this test
            svc.create_artifact(f"file{i}.txt", f"c{i}", "txt", "d", [], "local")

        svc._rate_limits.clear()
        with pytest.raises(PermissionError, match="quota"):
            svc.create_artifact("file100.txt", "extra", "txt", "d", [], "local")


class TestPathTraversal:
    def test_artifact_id_traversal_rejected(self, svc):
        with pytest.raises(ValueError, match="Invalid artifact ID"):
            svc._get_artifact_dir("../../../etc/passwd")

    def test_artifact_id_traversal_with_dotdot(self, svc):
        with pytest.raises(ValueError, match="Invalid artifact ID"):
            svc._get_artifact_dir("foo/../bar")

    def test_filename_never_in_path(self, svc):
        """Verify the content file is content.txt, not the user's filename."""
        a = svc.create_artifact("user-chosen-name.txt", "hello", "txt", "desc", [], "local")
        artifact_dir = svc._get_artifact_dir(a.id)
        # The content file should be content.txt
        assert (artifact_dir / "content.txt").exists()
        # The user's filename should NOT be a file
        assert not (artifact_dir / "user-chosen-name.txt").exists()


def test_tool_format_param_schema():
    from app.services.allai_tools import ALLAI_TOOLS

    tool = next(t for t in ALLAI_TOOLS if t["name"] == "create_artifact_from_query")
    schema = tool["input_schema"]

    assert schema["properties"]["format"]["enum"] == ["csv", "xlsx", "json", "parquet"]
    assert "format" in schema["required"]


class _FakeResult:
    def __init__(self, rows):
        self.description = [("id",), ("name",), ("amount",)]
        self._rows = list(rows)
        self._offset = 0

    def fetchmany(self, size):
        batch = self._rows[self._offset:self._offset + size]
        self._offset += len(batch)
        return batch


class _FakeConnection:
    def __init__(self, rows):
        self._rows = rows

    def execute(self, query):
        return _FakeResult(self._rows)

    def close(self):
        pass


class _FakeDuckDBService:
    def __init__(self, rows):
        self._rows = rows

    def create_ephemeral_connection(self):
        return _FakeConnection(self._rows)


class _FakeDuckDBContext:
    def __init__(self, rows):
        self._rows = rows

    def __enter__(self):
        return _FakeDuckDBService(self._rows)

    def __exit__(self, exc_type, exc, tb):
        return False


def _install_query_export_fakes(monkeypatch, tmp_path, rows):
    from app.config import settings
    from app.services import artifacts_service
    from app.services.artifacts_service import ArtifactsService

    export_dir = tmp_path / "export"
    export_dir.mkdir(parents=True)
    monkeypatch.setattr(artifacts_service, "EXPORT_DIR", export_dir)
    monkeypatch.setenv("HOST_EXPORT_DIR", str(export_dir))
    monkeypatch.setattr(settings, "mode", "standalone")

    class _SQLService:
        def validate_query(self, query):
            return True, None

        def _resolve_datasets(self, dataset_id):
            return []

    monkeypatch.setattr("app.services.sql_service.get_sql_service", lambda: _SQLService())
    monkeypatch.setattr("app.services.sql_service.SQLService._create_views", lambda conn, datasets: None)
    monkeypatch.setattr(
        "app.services.processing_service.get_processing_service",
        lambda: SimpleNamespace(list_datasets=lambda: [SimpleNamespace(id="ds")]),
    )
    monkeypatch.setattr(
        "app.services.duckdb_service.ephemeral_duckdb_service",
        lambda: _FakeDuckDBContext(rows),
    )

    svc = ArtifactsService()
    svc._artifacts_dir = tmp_path / "artifacts"
    svc._artifacts_dir.mkdir(parents=True)
    svc._rate_limits.clear()
    return svc, export_dir


def _create_query_export(monkeypatch, tmp_path, fmt, rows=None, filename=None):
    rows = rows or [(1, "alpha", 10.5), (2, "beta", 20.25), (3, "gamma", 30)]
    svc, export_dir = _install_query_export_fakes(monkeypatch, tmp_path, rows)
    artifact = svc.create_artifact_from_query(
        filename=filename or f"query-export.{fmt}",
        query="SELECT * FROM dataset_ds",
        description="query export",
        user_id="local",
        fmt=fmt,
    )
    return svc, export_dir, artifact


def _read_export_rows(path, fmt):
    if fmt == "csv":
        import csv
        with open(path, newline="", encoding="utf-8") as f:
            return list(csv.DictReader(f))

    if fmt == "xlsx":
        from openpyxl import load_workbook
        wb = load_workbook(path)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        header = rows[0]
        return [dict(zip(header, row)) for row in rows[1:]]

    if fmt == "json":
        return json.loads(path.read_text(encoding="utf-8"))

    if fmt == "parquet":
        import pyarrow.parquet as pq
        return pq.read_table(path).to_pylist()

    raise AssertionError(f"unexpected format: {fmt}")


def test_standalone_writes_to_data_export(monkeypatch, tmp_path):
    svc, export_dir, artifact = _create_query_export(monkeypatch, tmp_path, "csv")

    export_path = export_dir / "query-export.csv"
    assert export_path.exists()
    assert artifact.format == "csv"
    assert (svc._get_artifact_dir(artifact.id) / "content.csv").exists()


def test_export_csv_format(monkeypatch, tmp_path):
    _, export_dir, _ = _create_query_export(monkeypatch, tmp_path, "csv")

    rows = _read_export_rows(export_dir / "query-export.csv", "csv")
    assert rows[0]["name"] == "alpha"
    assert rows[1]["amount"] == "20.25"


def test_export_xlsx_format(monkeypatch, tmp_path):
    _, export_dir, _ = _create_query_export(monkeypatch, tmp_path, "xlsx")

    rows = _read_export_rows(export_dir / "query-export.xlsx", "xlsx")
    assert rows[0]["name"] == "alpha"
    assert rows[2]["amount"] == 30


def test_export_json_format(monkeypatch, tmp_path):
    _, export_dir, _ = _create_query_export(monkeypatch, tmp_path, "json")

    rows = _read_export_rows(export_dir / "query-export.json", "json")
    assert rows == [
        {"id": 1, "name": "alpha", "amount": 10.5},
        {"id": 2, "name": "beta", "amount": 20.25},
        {"id": 3, "name": "gamma", "amount": 30},
    ]


def test_export_parquet_format(monkeypatch, tmp_path):
    _, export_dir, _ = _create_query_export(monkeypatch, tmp_path, "parquet")

    rows = _read_export_rows(export_dir / "query-export.parquet", "parquet")
    assert rows[0]["name"] == "alpha"
    assert rows[2]["amount"] == 30


def test_path_canonicalization_blocks_symlink_and_traversal(monkeypatch, tmp_path):
    from app.services import artifacts_service
    from app.services.artifacts_service import ArtifactsService

    export_dir = tmp_path / "export"
    export_dir.mkdir()
    monkeypatch.setattr(artifacts_service, "EXPORT_DIR", export_dir)
    monkeypatch.setenv("HOST_EXPORT_DIR", str(export_dir))

    assert ArtifactsService._canonicalize_export_destination("safe.csv", "csv") == (
        export_dir / "safe.csv"
    ).resolve(strict=False)

    with pytest.raises(ValueError, match="path separators"):
        ArtifactsService._canonicalize_export_destination("../escape.csv", "csv")

    symlink_path = export_dir / "linked.csv"
    symlink_path.symlink_to(tmp_path / "outside.csv")
    with pytest.raises(ValueError, match="symlink|escapes"):
        ArtifactsService._canonicalize_export_destination("linked.csv", "csv")


def test_export_row_count_parity(monkeypatch, tmp_path):
    expected_count = 3
    for fmt in ["csv", "xlsx", "json", "parquet"]:
        _, export_dir, _ = _create_query_export(
            monkeypatch,
            tmp_path / fmt,
            fmt,
            filename=f"query-export-{fmt}.{fmt}",
        )
        rows = _read_export_rows(export_dir / f"query-export-{fmt}.{fmt}", fmt)
        assert len(rows) == expected_count


def test_call_path_gating_ownership_split(monkeypatch, tmp_path):
    from app.config import settings

    svc, _ = _install_query_export_fakes(monkeypatch, tmp_path, [(1, "alpha", 10)])
    monkeypatch.setattr(settings, "mode", "connected")

    with pytest.raises(PermissionError, match="standalone"):
        svc.create_artifact_from_query(
            filename="query-export.csv",
            query="SELECT * FROM dataset_ds",
            description="query export",
            user_id="local",
        )
